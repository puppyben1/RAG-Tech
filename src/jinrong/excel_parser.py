from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter

from .utils import norm_text, to_number


@dataclass(frozen=True)
class CellFact:
    sheet_name: str
    cell_ref: str
    row_index: int
    col_index: int
    row_header: str
    col_header: str
    value_raw: Any
    value_num: float | None
    unit: str | None
    table_title: str | None


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_unit(rows: list[list[Any]]) -> str | None:
    for row in rows[:8]:
        for value in row:
            text = _stringify(value)
            if "单位" in text:
                return text.replace("：", ":")
    return None


def _find_title(rows: list[list[Any]]) -> str | None:
    for row in rows[:5]:
        values = [_stringify(v) for v in row if _stringify(v)]
        if values and "单位" not in values[0]:
            return values[0]
    return None


def _detect_header_row(rows: list[list[Any]]) -> int:
    for i, row in enumerate(rows[:10]):
        if any(norm_text(v) == "项目" for v in row):
            return i
    best_idx = 0
    best_score = -1
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for v in row if _stringify(v))
        numeric = sum(1 for v in row if to_number(v) is not None)
        text_score = non_empty - numeric
        if non_empty >= 2 and text_score > best_score:
            best_idx = i
            best_score = text_score
    return best_idx


def _left_header(row: list[Any], row_idx: int, header_col_limit: int) -> str:
    candidates = []
    for i in range(min(header_col_limit, len(row))):
        text = _stringify(row[i])
        if text and to_number(text) is None:
            candidates.append(text)
    if candidates:
        return " ".join(candidates)
    return f"第{row_idx}行"


def _column_header(rows: list[list[Any]], header_row_idx: int, col_idx: int) -> str:
    pieces: list[str] = []
    for r in range(0, header_row_idx + 1):
        row = rows[r]
        text = ""
        if col_idx < len(row):
            text = _stringify(row[col_idx])
        if not text and col_idx > 0:
            for left in range(col_idx - 1, 0, -1):
                if left < len(row):
                    text = _stringify(row[left])
                    if text:
                        break
        if text:
            if text and "单位" not in text:
                pieces.append(text)
    seen: list[str] = []
    for piece in pieces:
        if piece not in seen:
            seen.append(piece)
    return "-".join(seen)


def _trim_rows(rows: list[list[Any]], max_cols_cap: int = 128) -> list[list[Any]]:
    last_col = 0
    for row in rows:
        for idx, value in enumerate(row, start=1):
            if _stringify(value):
                last_col = max(last_col, idx)
    last_col = min(last_col, max_cols_cap)
    return [row[:last_col] for row in rows]


@lru_cache(maxsize=128)
def parse_excel(path_str: str) -> tuple[CellFact, ...]:
    path = Path(path_str)
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return parse_xlsx(path_str)
    if ext == ".xls":
        return parse_xls(path_str)
    raise ValueError(f"unsupported excel extension: {path.suffix}")


@lru_cache(maxsize=128)
def parse_xlsx(path_str: str) -> tuple[CellFact, ...]:
    path = Path(path_str)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    facts: list[CellFact] = []
    for ws in wb.worksheets:
        rows = _trim_rows([list(row) for row in ws.iter_rows(values_only=True)])
        if not rows:
            continue
        unit = _find_unit(rows)
        title = _find_title(rows)
        header_row_idx = _detect_header_row(rows)
        header_col_limit = 2
        max_cols = max(len(row) for row in rows)
        col_headers = [_column_header(rows, header_row_idx, c) for c in range(max_cols)]
        for r_idx, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
            row_header = _left_header(row, r_idx, header_col_limit)
            for c_idx, value in enumerate(row, start=1):
                value_num = to_number(value)
                if value_num is None:
                    continue
                col_header = col_headers[c_idx - 1] if c_idx - 1 < len(col_headers) else get_column_letter(c_idx)
                facts.append(
                    CellFact(
                        sheet_name=ws.title,
                        cell_ref=f"{get_column_letter(c_idx)}{r_idx}",
                        row_index=r_idx,
                        col_index=c_idx,
                        row_header=row_header,
                        col_header=col_header,
                        value_raw=value,
                        value_num=value_num,
                        unit=unit,
                        table_title=title,
                    )
                )
    return tuple(facts)


@lru_cache(maxsize=128)
def parse_xls(path_str: str) -> tuple[CellFact, ...]:
    path = Path(path_str)
    wb = xlrd.open_workbook(path)
    facts: list[CellFact] = []
    for ws in wb.sheets():
        rows = _trim_rows([ws.row_values(r_idx) for r_idx in range(ws.nrows)])
        if not rows:
            continue
        unit = _find_unit(rows)
        title = _find_title(rows)
        header_row_idx = _detect_header_row(rows)
        header_col_limit = 2
        max_cols = max(len(row) for row in rows)
        col_headers = [_column_header(rows, header_row_idx, c) for c in range(max_cols)]
        for r_idx, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
            row_header = _left_header(row, r_idx, header_col_limit)
            for c_idx, value in enumerate(row, start=1):
                value_num = to_number(value)
                if value_num is None:
                    continue
                col_header = col_headers[c_idx - 1] if c_idx - 1 < len(col_headers) else get_column_letter(c_idx)
                facts.append(
                    CellFact(
                        sheet_name=ws.name,
                        cell_ref=f"{get_column_letter(c_idx)}{r_idx}",
                        row_index=r_idx,
                        col_index=c_idx,
                        row_header=row_header,
                        col_header=col_header,
                        value_raw=value,
                        value_num=value_num,
                        unit=unit,
                        table_title=title,
                    )
                )
    return tuple(facts)


def filter_sheet(facts: list[CellFact] | tuple[CellFact, ...], sheet_name: str | None) -> list[CellFact]:
    if not sheet_name:
        return list(facts)
    target = norm_text(sheet_name)
    matched = [f for f in facts if target and (target in norm_text(f.sheet_name) or norm_text(f.sheet_name) in target)]
    return matched or list(facts)
