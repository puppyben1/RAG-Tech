from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from .config import WENDANG_DIR


@dataclass
class QAItem:
    id: str
    source_type: str
    difficulty: str
    difficulty_cn: str
    qa_type: str
    question: str
    options: dict[str, Any]
    answer: str
    answer_text: Any
    evidence: str
    source_title: str
    file_label: str


def find_qa_path(root: Path = WENDANG_DIR) -> Path:
    matches = list(root.glob("**/QA数据.xlsx"))
    if not matches:
        raise FileNotFoundError(f"未找到 QA数据.xlsx: {root}")
    return matches[0]


def load_qa(path: Path | None = None) -> list[QAItem]:
    qa_path = path or find_qa_path()
    wb = openpyxl.load_workbook(qa_path, data_only=True, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows: list[QAItem] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        rows.append(
            QAItem(
                id=str(row["id"]),
                source_type=str(row["source_type"]),
                difficulty=str(row["difficulty"]),
                difficulty_cn=str(row["difficulty_cn"]),
                qa_type=str(row["qa_type"]),
                question=str(row["question"]),
                options={
                    "A": row["option_a"],
                    "B": row["option_b"],
                    "C": row["option_c"],
                    "D": row["option_d"],
                },
                answer=str(row["answer"]),
                answer_text=row["answer_text"],
                evidence=str(row["evidence"]),
                source_title=str(row["source_title"]),
                file_label=str(row["file_label"]),
            )
        )
    return rows

