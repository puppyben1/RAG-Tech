from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import MANIFEST_PATH, PROJECT_ROOT
from .qa_data import QAItem, load_qa
from .qa_data_audit import EVIDENCE_PATH, find_qa_candidates
from .utils import ensure_dir, nearly_equal, norm_text, read_jsonl, write_jsonl


REVIEW_FIELDS = (
    "id",
    "migration_status",
    "source_type",
    "source_title",
    "file_label",
    "candidate_doc_ids",
    "candidate_file_names",
    "original_evidence",
    "selected_doc_id",
    "page_no",
    "article_no",
    "sheet_name",
    "cell_ref",
    "review_notes",
)


def migrate_qa_data(
    qa_path: Path,
    output_path: Path,
    review_path: Path,
    manifest_path: Path = MANIFEST_PATH,
    project_root: Path = PROJECT_ROOT,
) -> dict[str, Any]:
    if output_path.resolve() == qa_path.resolve() or review_path.resolve() == qa_path.resolve():
        raise ValueError("migration outputs must not overwrite the source QA workbook")

    items = load_qa(qa_path)
    manifest = read_jsonl(manifest_path)
    by_file_name = {str(row.get("file_name")): row for row in manifest if row.get("file_name")}
    migrated: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []

    for item in items:
        evidence_file_name = _evidence_file_name(item.evidence)
        candidates = find_qa_candidates(item.file_label, item.source_title, evidence_file_name, manifest, by_file_name)
        locator_specs = _locator_specs(item)
        locators, locator_valid = _validate_locators(candidates, locator_specs, project_root)
        row = _migration_row(item, candidates, locators, locator_valid, project_root)
        migrated.append(row)
        if row["migration_status"] != "ready_candidate":
            review_rows.append(_review_row(item, candidates, row["migration_status"], locators))

    write_jsonl(output_path, migrated)
    _write_review_csv(review_path, review_rows)
    counts = Counter(row["migration_status"] for row in migrated)
    return {
        "status": "review_required" if review_rows else "ready_for_reviewed_freeze",
        "source_qa": qa_path.as_posix(),
        "manifest": manifest_path.as_posix(),
        "candidate_output": output_path.as_posix(),
        "review_worklist": review_path.as_posix(),
        "total": len(migrated),
        "ready_candidates": counts["ready_candidate"],
        "document_matched_locator_missing": counts["document_matched_locator_missing"],
        "locator_validation_failed": counts["locator_validation_failed"],
        "ambiguous_document": counts["ambiguous_document"],
        "unresolved_document": counts["unresolved_document"],
        "review_rows": len(review_rows),
    }


def _migration_row(
    item: QAItem,
    candidates: list[dict[str, Any]],
    locators: list[dict[str, Any]],
    locator_valid: bool,
    project_root: Path,
) -> dict[str, Any]:
    matched = candidates[0] if len(candidates) == 1 else None
    if not candidates:
        status = "unresolved_document"
    elif len(candidates) > 1:
        status = "ambiguous_document"
    elif not locators:
        status = "document_matched_locator_missing"
    elif not locator_valid:
        status = "locator_validation_failed"
    else:
        status = "ready_candidate"

    doc_id = str(matched.get("doc_id")) if matched else None
    local_path = _project_local_path(matched, project_root) if matched else None
    gold_evidence = [{"doc_id": doc_id, **locator} for locator in locators] if doc_id else []
    return {
        "id": item.id,
        "migration_status": status,
        "source_type": item.source_type,
        "difficulty": item.difficulty,
        "qa_type": item.qa_type,
        "question": item.question,
        "options": item.options,
        "answer": item.answer,
        "answer_text": item.answer_text,
        "original_evidence": item.evidence,
        "source_title": item.source_title,
        "file_label": item.file_label,
        "doc_id": doc_id,
        "expected_doc_ids": [doc_id] if doc_id else [],
        "expected_evidence_type": "table_cell" if item.source_type == "excel" else "text_unit",
        "gold_evidence": gold_evidence,
        "local_path": local_path,
        "source_url": matched.get("source_url") if matched else None,
        "tags": [item.source_type, item.qa_type, item.difficulty],
        "candidate_doc_ids": [str(row.get("doc_id")) for row in candidates],
    }


def _project_local_path(row: dict[str, Any], project_root: Path) -> str | None:
    file_name = str(row.get("file_name") or "")
    candidate = project_root / "wendang" / "data" / file_name
    return f"wendang/data/{file_name}" if file_name and candidate.is_file() else None


def _locator_specs(item: QAItem) -> list[dict[str, Any]]:
    if item.source_type != "excel":
        return []
    sheet = re.search(r"工作表[：:]\s*([^；;]+)", item.evidence)
    if not sheet:
        return []
    sheet_name = sheet.group(1).strip()
    specs: list[dict[str, Any]] = []
    single_cell = re.search(r"单元格[：:]\s*([A-Za-z]+\d+)", item.evidence)
    raw_value = re.search(r"原始值[：:]\s*([^；;。]+)", item.evidence)
    if single_cell:
        specs.append({"sheet_name": sheet_name, "cell_ref": single_cell.group(1).upper(), "expected_value": raw_value.group(1).strip() if raw_value else None})
    for expected_value, cell_ref in re.findall(r"=([^；;，,()]+)\(([A-Za-z]+\d+)\)", item.evidence):
        specs.append({"sheet_name": sheet_name, "cell_ref": cell_ref.upper(), "expected_value": expected_value.strip()})
    return list({spec["cell_ref"]: spec for spec in specs}.values())


def _validate_locators(
    candidates: list[dict[str, Any]],
    specs: list[dict[str, Any]],
    project_root: Path,
) -> tuple[list[dict[str, Any]], bool]:
    if len(candidates) != 1 or not specs:
        return [{"sheet_name": spec["sheet_name"], "cell_ref": spec["cell_ref"]} for spec in specs], False
    file_name = str(candidates[0].get("file_name") or "")
    path = project_root / "wendang" / "data" / file_name
    if path.suffix.lower() not in {".xlsx", ".xlsm"} or not path.is_file():
        return [], False
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return [], False
    try:
        normalized_sheets: dict[str, list[str]] = {}
        for sheet_name in workbook.sheetnames:
            normalized_sheets.setdefault(norm_text(sheet_name), []).append(sheet_name)
        validated: list[dict[str, Any]] = []
        for spec in specs:
            matches = normalized_sheets.get(norm_text(spec["sheet_name"]), [])
            if len(matches) != 1:
                return validated, False
            actual_sheet = matches[0]
            actual_value = workbook[actual_sheet][spec["cell_ref"]].value
            if actual_value is None:
                return validated, False
            expected_value = spec.get("expected_value")
            if expected_value is not None and not nearly_equal(actual_value, expected_value):
                return validated, False
            validated.append({"sheet_name": actual_sheet, "cell_ref": spec["cell_ref"]})
        return validated, True
    finally:
        workbook.close()


def _evidence_file_name(evidence: str) -> str | None:
    match = EVIDENCE_PATH.search(evidence.replace("\\", "/"))
    return match.group(1) if match else None


def _review_row(
    item: QAItem,
    candidates: list[dict[str, Any]],
    status: str,
    locators: list[dict[str, Any]],
) -> dict[str, str]:
    return {
        "id": item.id,
        "migration_status": status,
        "source_type": item.source_type,
        "source_title": item.source_title,
        "file_label": item.file_label,
        "candidate_doc_ids": ";".join(str(row.get("doc_id")) for row in candidates),
        "candidate_file_names": ";".join(str(row.get("file_name")) for row in candidates),
        "original_evidence": item.evidence,
        "selected_doc_id": "",
        "page_no": "",
        "article_no": "",
        "sheet_name": ";".join(dict.fromkeys(str(locator.get("sheet_name") or "") for locator in locators)),
        "cell_ref": ";".join(str(locator.get("cell_ref") or "") for locator in locators),
        "review_notes": "",
    }


def _write_review_csv(path: Path, rows: list[dict[str, str]]) -> None:
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
